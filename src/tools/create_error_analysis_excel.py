import os
import random
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import io

from typing import Optional

# ================= CẤU HÌNH TRỰC TIẾP =================
PREDICTIONS_FILE = "/home/laptq/laptq-fs26-shoplifting-detection/outputs/trivials/predict_2dcnn_action/predictions.txt"
VAL_DATA = {
    "cho_tay_vao_tui_quan": ["data/processed/fs26/action_recognition/classification/action.for_CNN.8_classes.manually_selected/cho_tay_vao_tui_quan/val.easy.txt"],
    "cho_tay_vao_tui_ao": ["data/processed/fs26/action_recognition/classification/action.for_CNN.8_classes.manually_selected/cho_tay_vao_tui_ao/val.easy.txt"],
    "cho_tay_vao_tui_deo_tren_nguoi": ["data/processed/fs26/action_recognition/classification/action.for_CNN.8_classes.manually_selected/cho_tay_vao_tui_deo_tren_nguoi/val.easy.txt"],
    "cho_tay_vao_tui_cam_tren_tay": ["data/processed/fs26/action_recognition/classification/action.for_CNN.8_classes.manually_selected/cho_tay_vao_tui_cam_tren_tay/val.easy.txt"]
}
OUTPUT_FILE = "data/tmp/error_analysis/action_recognition/classification/excels/error_analysis.xlsx"  # Path tới file excel kết quả
MAX_IMAGES_PER_SHEET = 5000  # Giới hạn số lượng ảnh lỗi trên mỗi sheet
SEED = 42  # None: sắp xếp theo tên; int: xáo trộn ngẫu nhiên theo seed
# =====================================================

def load_ground_truth(val_data: dict[str, list[str]]) -> dict[str, str]:
    gt_map: dict[str, str] = {}
    
    for class_name, file_list in val_data.items():
        for file_path in file_list:
            if not os.path.exists(file_path):
                print(f"Warning: File {file_path} not found.")
                continue
            with open(file_path, "r", encoding="utf-8") as tf:
                for line in tf:
                    img_path = line.strip()
                    if img_path:
                        normalized_path = os.path.normpath(img_path)
                        gt_map[normalized_path] = class_name
                        
    return gt_map

def main(
    predictions_file: str,
    val_data: dict[str, list[str]],
    output_file: str,
    max_images: int,
    seed_value: Optional[int]
) -> None:
    print("Loading ground truth data...")
    gt_map = load_ground_truth(val_data)
    print(f"Loaded {len(gt_map)} ground truth items.")
    
    if not os.path.exists(predictions_file):
        print(f"Error: Predictions file {predictions_file} does not exist.")
        return
        
    print(f"Reading predictions from: {predictions_file}")
    
    # Store errors grouped by ground truth class
    errors_by_class: dict[str, list[dict[str, str]]] = {}
    
    with open(predictions_file, "r", encoding="utf-8") as f:
        for line in f:
            parts = line.strip().split("\t")
            if len(parts) < 3:
                continue
            
            img_path = parts[0]
            pred_label = parts[1]
            pred_score = parts[2]
            
            normalized_path = os.path.normpath(img_path)
            
            if normalized_path in gt_map:
                true_label = gt_map[normalized_path]
                if true_label != pred_label:
                    if true_label not in errors_by_class:
                        errors_by_class[true_label] = []
                    
                    errors_by_class[true_label].append({
                        "image_path": normalized_path,
                        "true_label": true_label,
                        "pred_label": pred_label,
                        "score": pred_score
                    })
                    
    # Initialize openpyxl workbook
    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)
        
    for class_name, errors in errors_by_class.items():
        if seed_value is None:
            errors.sort(key=lambda x: x["image_path"])
        else:
            rng = random.Random(seed_value)
            rng.shuffle(errors)
            
        print(f"Processing class: {class_name} with {len(errors)} errors.")
        ws = wb.create_sheet(title=class_name)
        
        # Write headers
        headers = ["ảnh", "nhãn đúng", "nhãn dự đoán", "score dự đoán", "ghi chú", "đường dẫn ảnh"]
        ws.append(headers)
        
        # Configure columns widths
        ws.column_dimensions["A"].width = 20  # For image
        ws.column_dimensions["B"].width = 30  # True label
        ws.column_dimensions["C"].width = 30  # Pred label
        ws.column_dimensions["D"].width = 15  # Score
        ws.column_dimensions["E"].width = 25  # Notes
        ws.column_dimensions["F"].width = 80  # Image path
        
        # Add details and insert images
        # We start writing from row 2 (row 1 is header)
        for idx, error in enumerate(errors[:max_images]):
            row_idx = idx + 2
            ws.row_dimensions[row_idx].height = 90  # Height in points
            
            # Write text columns
            ws.cell(row=row_idx, column=2, value=error["true_label"])
            ws.cell(row=row_idx, column=3, value=error["pred_label"])
            ws.cell(row=row_idx, column=4, value=float(error["score"]))
            ws.cell(row=row_idx, column=5, value="")  # Empty note
            ws.cell(row=row_idx, column=6, value=error["image_path"])
            
            # Insert image
            img_path = error["image_path"]
            if os.path.exists(img_path):
                try:
                    pil_img = PILImage.open(img_path)
                    # Resize while keeping aspect ratio, maximum dimensions 140x110
                    pil_img.thumbnail((140, 110))
                    
                    img_byte_arr = io.BytesIO()
                    # Convert to RGB to ensure JPEG compatibility
                    if pil_img.mode in ("RGBA", "P"):
                        pil_img = pil_img.convert("RGB")
                    pil_img.save(img_byte_arr, format="JPEG", quality=85)
                    img_byte_arr.seek(0)
                    
                    ox_img = OpenpyxlImage(img_byte_arr)
                    # Add image to cell A{row_idx}
                    ws.add_image(ox_img, f"A{row_idx}")
                except Exception as e:
                    ws.cell(row=row_idx, column=1, value=f"Error loading image: {str(e)}")
            else:
                ws.cell(row=row_idx, column=1, value="File not found")
                
    wb.save(output_file)
    print(f"Saved error analysis excel file to {output_file}")

if __name__ == "__main__":
    main(
        predictions_file=PREDICTIONS_FILE,
        val_data=VAL_DATA,
        output_file=OUTPUT_FILE,
        max_images=MAX_IMAGES_PER_SHEET,
        seed_value=SEED
    )
