import argparse
import os
import shutil
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_INPUT_FILE: str = (
    "/home/laptq/classification/outputs/excels/error_analysis_annot.xlsx"
)

DEFAULT_IMAGE_OUTPUT_DIR: str = (
    "/home/laptq/classification/outputs/categorized_images"
)

PRIORITY_ORDER: List[str] = ["easy", "medium", "hard"]

PATH_SRC_PREFIX: str = (
    "/home/laptq/laptq-fs26-shoplifting-detection/outputs/"
    "helper--extract--crops--from--detection/fs26"
)

PATH_DST_PREFIX: str = (
    "/home/laptq/classification_multilabel/externals/classification/outputs/"
    "copy_images"
)


def sort_key(category_name: str) -> Tuple[int, str]:
    if category_name in PRIORITY_ORDER:
        return (PRIORITY_ORDER.index(category_name), category_name)
    else:
        return (len(PRIORITY_ORDER), category_name)


def extract_categories_from_sheet(sheet: Worksheet) -> Dict[str, int]:
    categories: Dict[str, int] = {}
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if len(row) < 5:
            continue

        true_label: Any = row[1].value
        if true_label is None:
            continue

        cell_val: Any = row[4].value
        # Only count if the category is annotated (not None or empty)
        if cell_val is None:
            continue

        val_str: str = str(cell_val).strip()
        if val_str == "":
            continue

        if val_str not in categories:
            categories[val_str] = 0
        categories[val_str] += 1
    return categories


def copy_annotated_images(
    sheet: Worksheet,
    sheet_name: str,
    output_dir: str
) -> None:
    category_counters: Dict[str, int] = {}

    for row in sheet.iter_rows(min_row=2, values_only=False):
        if len(row) < 6:
            continue

        true_label: Any = row[1].value
        if true_label is None:
            continue

        category_val: Any = row[4].value
        if category_val is None:
            continue
        category_str: str = str(category_val).strip()
        if category_str == "":
            continue

        img_path_val: Any = row[5].value
        if img_path_val is None:
            continue
        img_path: str = str(img_path_val).strip()
        if img_path == "":
            continue

        # Replace beginning of image path
        replaced_path: str = img_path.replace(PATH_SRC_PREFIX, PATH_DST_PREFIX)

        if not os.path.exists(replaced_path):
            print(f"Warning: File not found at {replaced_path}")
            continue

        if category_str not in category_counters:
            category_counters[category_str] = 1
        else:
            category_counters[category_str] += 1

        count: int = category_counters[category_str]

        _, ext = os.path.splitext(replaced_path)
        if not ext:
            ext = ".jpg"

        dest_dir: str = os.path.join(output_dir, sheet_name, category_str)
        os.makedirs(dest_dir, exist_ok=True)

        dest_filename: str = f"{count}{ext}"
        dest_path: str = os.path.join(dest_dir, dest_filename)

        try:
            shutil.copy2(replaced_path, dest_path)
        except Exception as e:
            print(f"Error copying {replaced_path} to {dest_path}: {e}")


def print_unified_text_table(
    sheet_names: List[str],
    all_categories: List[str],
    sheet_data: Dict[str, Dict[str, int]],
    sheet_totals: Dict[str, int],
    category_totals: Dict[str, int],
    grand_total: int
) -> None:
    headers = ["Nhãn đúng (Ground Truth)"] + all_categories + ["Tổng số đã gán"]

    col_widths: List[int] = [35]
    for _ in all_categories:
        col_widths.append(15)
    col_widths.append(18)

    def make_separator() -> str:
        parts = []
        for w in col_widths:
            parts.append("-" * w)
        return "+".join(parts)

    separator = make_separator()

    print(separator)
    header_parts = []
    for h, w in zip(headers, col_widths):
        if w == 35:
            header_parts.append(f"{h:<{w}}")
        else:
            header_parts.append(f"{h:^{w}}")
    print("|".join(header_parts))
    print(separator)

    # Data rows
    for name in sheet_names:
        row_parts = []
        row_parts.append(f"{name:<35}")

        s_total = sheet_totals[name]
        s_cats = sheet_data[name]

        for cat in all_categories:
            if cat in s_cats:
                count = s_cats[cat]
            else:
                count = 0
            pct = (count / s_total * 100.0) if s_total > 0 else 0.0
            val_str = f"{pct:.1f}% ({count})"
            row_parts.append(f"{val_str:^15}")

        tot_val_str = f"100.0% ({s_total})"
        row_parts.append(f"{tot_val_str:^18}")
        print("|".join(row_parts))

    print(separator)

    # Total row
    tot_parts = []
    tot_parts.append(f"{'Tổng cộng':<35}")
    for cat in all_categories:
        count = category_totals[cat]
        pct = (count / grand_total * 100.0) if grand_total > 0 else 0.0
        val_str = f"{pct:.1f}% ({count})"
        tot_parts.append(f"{val_str:^15}")

    grand_val_str = f"100.0% ({grand_total})"
    tot_parts.append(f"{grand_val_str:^18}")
    print("|".join(tot_parts))
    print(separator)
    print()


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Calculate statistics of category annotations "
            "and copy annotated image crops."
        )
    )
    parser.add_argument(
        "--input",
        type=str,
        default=DEFAULT_INPUT_FILE,
        help=f"Path to the annotated excel file (default: {DEFAULT_INPUT_FILE})"
    )
    parser.add_argument(
        "--image-out",
        type=str,
        default=DEFAULT_IMAGE_OUTPUT_DIR,
        help=f"Directory to copy categorized images to (default: {DEFAULT_IMAGE_OUTPUT_DIR})"
    )

    args = parser.parse_args()
    input_file: str = args.input
    image_out_dir: str = args.image_out

    if not os.path.exists(input_file):
        print(f"Error: Input file {input_file} does not exist.")
        return

    print(f"Loading workbook (read-only): {input_file}")
    wb = openpyxl.load_workbook(input_file, read_only=True)

    # 1. Collect all categories and compute stats first
    all_categories_set = set()
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        cats = extract_categories_from_sheet(sheet)
        for cat in cats:
            all_categories_set.add(cat)

    all_categories: List[str] = sorted(list(all_categories_set), key=sort_key)

    sheet_data: Dict[str, Dict[str, int]] = {}
    sheet_totals: Dict[str, int] = {}
    category_totals: Dict[str, int] = {}
    grand_total: int = 0

    for cat in all_categories:
        category_totals[cat] = 0

    sheet_names: List[str] = wb.sheetnames

    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        cats = extract_categories_from_sheet(sheet)

        sheet_data[sheet_name] = cats
        total = sum(cats.values())
        sheet_totals[sheet_name] = total
        grand_total += total

        for cat, count in cats.items():
            category_totals[cat] += count

    # 2. Print console statistics
    print_unified_text_table(
        sheet_names=sheet_names,
        all_categories=all_categories,
        sheet_data=sheet_data,
        sheet_totals=sheet_totals,
        category_totals=category_totals,
        grand_total=grand_total
    )

    # 3. Copy annotated images to output folders
    print(f"Copying categorized images to: {image_out_dir}")
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        copy_annotated_images(sheet, sheet_name, image_out_dir)
    print("Done!")


if __name__ == "__main__":
    main()
